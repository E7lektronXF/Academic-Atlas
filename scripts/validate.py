"""
Validates database/AcademicAtlas.xlsx against docs/DatabaseSchema.md.

This is the automated guardrail behind the project's "AI-maintainable,
consistent formatting" promise: it checks IDs, enums, required fields, dates,
and URLs so that malformed data is caught before it reaches the website.

Run it after editing the spreadsheet and before committing:

    python scripts/validate.py

Exit code is 0 when every record is valid, and 1 when any error is found
(so it can gate contributions in CI). Warnings never fail the build.
"""

import re
import sys
from datetime import date
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = REPO_ROOT / "database" / "AcademicAtlas.xlsx"

# Column order must match docs/DatabaseSchema.md exactly.
EXPECTED_HEADERS = [
    "ID",
    "Category",
    "Name",
    "Organizer",
    "Country_Region",
    "Eligibility",
    "Description",
    "Official_URL",
    "Application_Deadline",
    "Event_Dates",
    "Cost",
    "Format",
    "Status",
    "Last_Verified",
    "Notes",
    "Eligibility_Scope",
    "Qualifies_For",
    "Subject",
    "Typical_Window",
]

# Optional columns may legitimately hold nothing; everything else is required.
OPTIONAL_COLUMNS = {"Event_Dates", "Notes", "Qualifies_For", "Typical_Window"}

# Prefix -> full Category name (docs/DatabaseSchema.md identifier convention).
PREFIX_TO_CATEGORY = {
    "COMP": "International competitions",
    "RES": "Research programs / opportunities",
    "SUM": "Summer schools",
    "SCHOL": "Scholarships",
    "COURSE": "Academic courses",
    "INNOV": "Innovation challenges",
    "JOUR": "Academic journals & publications",
    "CONF": "Conferences & academic events",
}
VALID_CATEGORIES = set(PREFIX_TO_CATEGORY.values())
VALID_FORMATS = {"Online", "In-person", "Hybrid", "UNKNOWN"}
VALID_STATUSES = {"Active", "Upcoming", "Archived"}
VALID_SCOPES = {"International", "Türkiye only", "US only"}

# Controlled vocabulary for the multi-value Subject column (docs/DatabaseSchema.md
# -> Subject Vocabulary). Values are separated by ';' within a cell.
VALID_SUBJECTS = {
    "Mathematics",
    "Physics",
    "Chemistry",
    "Biology",
    "Computer Science",
    "Engineering & Robotics",
    "Earth & Space",
    "Environment",
    "Economics & Business",
    "Social Sciences & Humanities",
    "Linguistics",
    "Writing",
    "Arts & Design",
    "Medicine & Health",
    "Interdisciplinary",
}
WINDOW_MAX_LEN = 40  # Typical_Window is free text but should stay short.

ID_RE = re.compile(r"^(" + "|".join(PREFIX_TO_CATEGORY) + r")-(\d{4})$")
ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _is_iso_date(value):
    """True if value is a real YYYY-MM-DD calendar date."""
    if not ISO_DATE_RE.match(value):
        return False
    try:
        date.fromisoformat(value)
        return True
    except ValueError:
        return False


def _cell_str(value):
    """Normalize a cell to a stripped string (dates keep ISO form)."""
    if value is None:
        return ""
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def validate():
    errors = []
    warnings = []

    if not XLSX_PATH.exists():
        return [f"Database file not found: {XLSX_PATH}"], []

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if "Database" not in wb.sheetnames:
        return ['Missing required sheet "Database".'], []

    ws = wb["Database"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ["Database sheet is empty (no header row)."], []

    headers = [_cell_str(h) for h in rows[0][: len(EXPECTED_HEADERS)]]
    if headers != EXPECTED_HEADERS:
        errors.append(
            "Header row does not match the schema.\n"
            f"  expected: {EXPECTED_HEADERS}\n"
            f"  found:    {headers}"
        )
        # Column positions are unreliable now; stop before row checks.
        return errors, warnings

    seen_ids = {}
    numbers_by_prefix = {}
    qualifies_refs = []  # (label, referenced_id) — checked for existence after the loop.

    for row_idx, raw in enumerate(rows[1:], start=2):
        row = {h: _cell_str(v) for h, v in zip(EXPECTED_HEADERS, raw)}
        if not any(row.values()):
            continue  # blank spacer row

        rid = row["ID"]
        label = f"row {row_idx} (ID={rid or '?'})"

        # Required fields present.
        for col in EXPECTED_HEADERS:
            if col in OPTIONAL_COLUMNS:
                continue
            if not row[col]:
                errors.append(f"{label}: required column '{col}' is empty.")

        # ID format, uniqueness, prefix<->category consistency.
        m = ID_RE.match(rid)
        if not m:
            errors.append(
                f"{label}: ID '{rid}' is malformed "
                "(expected PREFIX-NNNN, e.g. COMP-0001)."
            )
        else:
            prefix, number = m.group(1), m.group(2)
            if rid in seen_ids:
                errors.append(
                    f"{label}: duplicate ID, already used on row {seen_ids[rid]}."
                )
            else:
                seen_ids[rid] = row_idx
            numbers_by_prefix.setdefault(prefix, []).append(int(number))
            expected_cat = PREFIX_TO_CATEGORY[prefix]
            if row["Category"] and row["Category"] != expected_cat:
                errors.append(
                    f"{label}: ID prefix '{prefix}' implies category "
                    f"'{expected_cat}', but Category is '{row['Category']}'."
                )

        # Enums.
        if row["Category"] and row["Category"] not in VALID_CATEGORIES:
            errors.append(f"{label}: Category '{row['Category']}' is not a known category.")
        if row["Format"] and row["Format"] not in VALID_FORMATS:
            errors.append(
                f"{label}: Format '{row['Format']}' must be one of {sorted(VALID_FORMATS)}."
            )
        if row["Status"] and row["Status"] not in VALID_STATUSES:
            errors.append(
                f"{label}: Status '{row['Status']}' must be one of {sorted(VALID_STATUSES)}."
            )
        if row["Eligibility_Scope"] and row["Eligibility_Scope"] not in VALID_SCOPES:
            errors.append(
                f"{label}: Eligibility_Scope '{row['Eligibility_Scope']}' must be one of "
                f"{sorted(VALID_SCOPES)}."
            )

        # Application_Deadline: ISO date, or the allowed special values.
        deadline = row["Application_Deadline"]
        if deadline and deadline not in {"Rolling", "UNKNOWN"} and not _is_iso_date(deadline):
            errors.append(
                f"{label}: Application_Deadline '{deadline}' must be YYYY-MM-DD, "
                "'Rolling', or 'UNKNOWN'."
            )

        # Last_Verified: real ISO date, not in the future.
        verified = row["Last_Verified"]
        if verified:
            if not _is_iso_date(verified):
                errors.append(
                    f"{label}: Last_Verified '{verified}' must be a valid YYYY-MM-DD date."
                )
            elif date.fromisoformat(verified) > date.today():
                errors.append(
                    f"{label}: Last_Verified '{verified}' is in the future."
                )

        # Official_URL scheme.
        url = row["Official_URL"]
        if url and not re.match(r"^https?://", url, re.IGNORECASE):
            errors.append(
                f"{label}: Official_URL '{url}' must start with http:// or https://."
            )

        # Qualifies_For (optional): must be a well-formed ID that isn't self-referential.
        qualifies = row["Qualifies_For"]
        if qualifies:
            if not ID_RE.match(qualifies):
                errors.append(
                    f"{label}: Qualifies_For '{qualifies}' must be a valid ID "
                    "(PREFIX-NNNN) referencing an existing record."
                )
            elif qualifies == rid:
                errors.append(f"{label}: Qualifies_For cannot reference the record's own ID.")
            else:
                qualifies_refs.append((label, qualifies))

        # Subject (required): one or more ';'-separated values from the vocabulary.
        # (An entirely empty Subject is already reported by the required-field check.)
        subject = row["Subject"]
        if subject:
            seen_subjects = set()
            for part in (p.strip() for p in subject.split(";")):
                if not part:
                    errors.append(
                        f"{label}: Subject has an empty value (check for a stray ';')."
                    )
                elif part not in VALID_SUBJECTS:
                    errors.append(
                        f"{label}: Subject '{part}' is not in the controlled vocabulary "
                        f"{sorted(VALID_SUBJECTS)}."
                    )
                elif part in seen_subjects:
                    errors.append(f"{label}: Subject '{part}' is listed more than once.")
                else:
                    seen_subjects.add(part)

        # Typical_Window (optional): free text, but kept short for the site pills.
        window = row["Typical_Window"]
        if window and len(window) > WINDOW_MAX_LEN:
            warnings.append(
                f"{label}: Typical_Window '{window}' is longer than "
                f"{WINDOW_MAX_LEN} characters."
            )

    # Qualifies_For must point at a record that actually exists (checked after the
    # full pass so a stage may appear before or after the final it feeds into).
    for label, ref in qualifies_refs:
        if ref not in seen_ids:
            errors.append(f"{label}: Qualifies_For '{ref}' does not match any record's ID.")

    # Sequential-ID gaps are a warning, not an error (records can be archived).
    for prefix, numbers in numbers_by_prefix.items():
        expected = set(range(1, max(numbers) + 1))
        missing = sorted(expected - set(numbers))
        if missing:
            gaps = ", ".join(f"{prefix}-{n:04d}" for n in missing)
            warnings.append(f"Non-sequential IDs for prefix '{prefix}': missing {gaps}.")

    return errors, warnings


def main():
    errors, warnings = validate()

    for w in warnings:
        print(f"WARNING: {w}")

    if errors:
        print(f"\nValidation FAILED with {len(errors)} error(s):\n")
        for e in errors:
            print(f"ERROR: {e}")
        sys.exit(1)

    suffix = f" ({len(warnings)} warning(s))" if warnings else ""
    print(f"Validation passed - no errors{suffix}.")


if __name__ == "__main__":
    main()
